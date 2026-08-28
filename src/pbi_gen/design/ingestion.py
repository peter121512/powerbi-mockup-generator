"""Data ingestion for Stage 13 — turns uploads / URLs / descriptions into a
provider-neutral `DataContext`.

Three input mechanisms, one output shape:
- `profile_spreadsheet(path)`  — CSV / XLSX upload
- `resolve_url(url)`           — externally hosted CSV/XLSX (abstraction; future
                                 connectors plug into the same path)
- `context_from_description()` — LLM-free heuristic inference from prose (a
                                 provider may enrich this, but a deterministic
                                 fallback always works offline/for tests)

Type/role inference is deterministic and dependency-light (stdlib csv + openpyxl
for xlsx). It does not build a full semantic model — just enough to ground design.
"""

from __future__ import annotations

import csv
import io
import re
import uuid
from pathlib import Path
from typing import Iterable, Optional

import requests

from .data_context import (
    DataContext,
    DataSource,
    FieldProfile,
    FieldRole,
    FieldType,
)

# ─────────────────────────────────────────────────────────────────────────────
# Type / role inference heuristics
# ─────────────────────────────────────────────────────────────────────────────

_CURRENCY_HINT = re.compile(r"(revenue|sales|cost|price|amount|arr|mrr|ltv|budget|profit|margin|spend|income|value)", re.I)
_PCT_HINT = re.compile(r"(rate|pct|percent|%|ratio|churn|retention|growth|share)", re.I)
_DATE_HINT = re.compile(r"(date|day|month|year|period|quarter|timestamp|_at$|_on$)", re.I)
_ID_HINT = re.compile(r"(^id$|_id$|code$|number$|guid|uuid|key$)", re.I)
_MEASURE_HINT = re.compile(r"(count|qty|quantity|total|sum|avg|units|orders|visits|sessions|volume)", re.I)

_DATE_VALUE = re.compile(r"^\d{4}[-/]\d{1,2}([-/]\d{1,2})?([ T]\d{1,2}:\d{2})?$|^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$")
_CURRENCY_VALUE = re.compile(r"^[£$€]\s?-?\d")
_PCT_VALUE = re.compile(r"^-?\d+(\.\d+)?\s?%$")
_INT_VALUE = re.compile(r"^-?\d{1,3}(,\d{3})*$|^-?\d+$")
_DEC_VALUE = re.compile(r"^-?\d*\.\d+$|^-?\d{1,3}(,\d{3})*\.\d+$")


def _infer_type(name: str, samples: list[str]) -> FieldType:
    non_null = [s.strip() for s in samples if s is not None and str(s).strip() != ""]
    if not non_null:
        # Fall back to the name hint.
        if _DATE_HINT.search(name):
            return FieldType.DATE
        if _PCT_HINT.search(name):
            return FieldType.PERCENTAGE
        if _CURRENCY_HINT.search(name):
            return FieldType.CURRENCY
        return FieldType.UNKNOWN

    def frac(pred) -> float:
        return sum(1 for v in non_null if pred(v)) / len(non_null)

    if frac(lambda v: bool(_DATE_VALUE.match(v))) >= 0.7 or _DATE_HINT.search(name):
        return FieldType.DATE
    if frac(lambda v: bool(_PCT_VALUE.match(v))) >= 0.6 or (_PCT_HINT.search(name) and frac(lambda v: bool(_DEC_VALUE.match(v) or _INT_VALUE.match(v))) >= 0.6):
        return FieldType.PERCENTAGE
    if frac(lambda v: bool(_CURRENCY_VALUE.match(v))) >= 0.5 or (_CURRENCY_HINT.search(name) and frac(lambda v: bool(_DEC_VALUE.match(v) or _INT_VALUE.match(v))) >= 0.6):
        return FieldType.CURRENCY
    if frac(lambda v: bool(_DEC_VALUE.match(v))) >= 0.6:
        return FieldType.DECIMAL
    if frac(lambda v: bool(_INT_VALUE.match(v))) >= 0.7:
        return FieldType.INTEGER
    if frac(lambda v: v.lower() in ("true", "false", "yes", "no", "y", "n")) >= 0.8:
        return FieldType.BOOLEAN
    return FieldType.STRING


def _infer_role(name: str, ftype: FieldType) -> FieldRole:
    if ftype in (FieldType.DATE, FieldType.DATETIME):
        return FieldRole.DATE
    if _ID_HINT.search(name):
        return FieldRole.IDENTIFIER
    if ftype in (FieldType.CURRENCY, FieldType.PERCENTAGE, FieldType.DECIMAL):
        return FieldRole.MEASURE
    if ftype == FieldType.INTEGER:
        # Integer + measure hint => measure, else likely dimension/identifier.
        if _MEASURE_HINT.search(name) or _CURRENCY_HINT.search(name):
            return FieldRole.MEASURE
        if _ID_HINT.search(name):
            return FieldRole.IDENTIFIER
        return FieldRole.MEASURE  # counts default to measure
    if ftype in (FieldType.STRING, FieldType.BOOLEAN):
        return FieldRole.DIMENSION
    return FieldRole.UNKNOWN


def _build_fields(entity: str, headers: list[str], rows: list[list[str]]) -> list[FieldProfile]:
    fields: list[FieldProfile] = []
    for ci, header in enumerate(headers):
        col_vals = [row[ci] for row in rows if ci < len(row)]
        samples = [str(v) for v in col_vals if v is not None and str(v).strip() != ""][:5]
        ftype = _infer_type(header, [str(v) for v in col_vals])
        role = _infer_role(header, ftype)
        distinct = len({str(v).strip() for v in col_vals if str(v).strip() != ""}) or None
        nulls = sum(1 for v in col_vals if v is None or str(v).strip() == "")
        null_ratio = (nulls / len(col_vals)) if col_vals else None
        fields.append(FieldProfile(
            name=header, entity=entity, field_type=ftype, role=role,
            sample_values=samples, distinct_count=distinct, null_ratio=null_ratio,
        ))
    return fields


def _assemble_context(sources: list[DataSource], all_fields: list[FieldProfile]) -> DataContext:
    entities = sorted({f.entity for f in all_fields if f.entity})
    measures = [f.name for f in all_fields if f.role == FieldRole.MEASURE]
    dims = [f.name for f in all_fields if f.role == FieldRole.DIMENSION]
    dates = [f.name for f in all_fields if f.role == FieldRole.DATE]

    assumptions: list[str] = []
    unknowns = [f.name for f in all_fields if f.field_type == FieldType.UNKNOWN]
    if unknowns:
        assumptions.append(f"Types inferred; unresolved for: {', '.join(unknowns[:5])}")
    if not dates:
        assumptions.append("No explicit date field detected; time-trend visuals may need a date dimension.")

    # Confidence: proportion of fields with a resolved type + presence of measures.
    resolved = sum(1 for f in all_fields if f.field_type != FieldType.UNKNOWN)
    type_conf = (resolved / len(all_fields)) if all_fields else 0.0
    has_measure = 1.0 if measures else 0.4
    confidence = round(min(1.0, 0.5 * type_conf + 0.5 * has_measure), 2)

    return DataContext(
        sources=sources,
        entities=entities,
        fields=all_fields,
        candidate_measures=measures,
        candidate_dimensions=dims,
        date_fields=dates,
        relationships=[],
        assumptions=assumptions,
        confidence=confidence,
        context_id=f"dc_{uuid.uuid4().hex[:10]}",
    )


# ─────────────────────────────────────────────────────────────────────────────
# CSV / XLSX profiling
# ─────────────────────────────────────────────────────────────────────────────

_MAX_SAMPLE_ROWS = 200


def _profile_csv_text(text: str, entity: str) -> tuple[list[FieldProfile], int]:
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], 0
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    sample = data[:_MAX_SAMPLE_ROWS]
    return _build_fields(entity, headers, sample), len(data)


def _profile_xlsx(path: Path) -> tuple[list[FieldProfile], list[str], int]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    all_fields: list[FieldProfile] = []
    total_rows = 0
    sheet_names: list[str] = []
    for ws in wb.worksheets:
        sheet_names.append(ws.title)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        data: list[list[str]] = []
        n = 0
        for r in rows_iter:
            n += 1
            if len(data) < _MAX_SAMPLE_ROWS:
                data.append(["" if c is None else str(c) for c in r])
        total_rows += n
        all_fields.extend(_build_fields(ws.title, headers, data))
    wb.close()
    return all_fields, sheet_names, total_rows


def profile_spreadsheet(path: str | Path) -> DataContext:
    """Profile a CSV or XLSX file into a DataContext."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {p}")
    ext = p.suffix.lower()
    if ext == ".csv":
        text = p.read_text(encoding="utf-8-sig", errors="replace")
        fields, rows = _profile_csv_text(text, p.stem)
        source = DataSource(name=p.name, kind="upload", location=str(p), row_count=rows,
                            sheet_names=[p.stem])
        return _assemble_context([source], fields)
    if ext in (".xlsx", ".xlsm"):
        fields, sheets, rows = _profile_xlsx(p)
        source = DataSource(name=p.name, kind="upload", location=str(p), row_count=rows,
                            sheet_names=sheets)
        return _assemble_context([source], fields)
    raise ValueError(f"Unsupported spreadsheet type: {ext} (supported: .csv, .xlsx)")


# ─────────────────────────────────────────────────────────────────────────────
# URL / file resolver abstraction
# ─────────────────────────────────────────────────────────────────────────────


class FileResolver:
    """Resolves an externally-hosted file reference into a DataContext.

    The abstraction is deliberately not local-upload-only: future connectors
    (auth'd sources, cloud stores) plug in by subclassing `fetch`.
    """

    def fetch(self, url: str) -> tuple[bytes, str]:
        """Return (content_bytes, suffix). Default: direct HTTP GET."""
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        # Derive suffix from URL or content-type.
        suffix = Path(url.split("?")[0]).suffix.lower()
        if not suffix:
            ctype = r.headers.get("Content-Type", "")
            if "csv" in ctype:
                suffix = ".csv"
            elif "spreadsheet" in ctype or "excel" in ctype:
                suffix = ".xlsx"
        return r.content, suffix


def resolve_url(url: str, resolver: Optional[FileResolver] = None) -> DataContext:
    """Resolve a URL to a DataContext using the same profiling path as uploads."""
    resolver = resolver or FileResolver()
    content, suffix = resolver.fetch(url)
    name = Path(url.split("?")[0]).name or "remote_file"
    if suffix == ".csv":
        fields, rows = _profile_csv_text(content.decode("utf-8-sig", errors="replace"), Path(name).stem)
        source = DataSource(name=name, kind="url", location=url, row_count=rows,
                            sheet_names=[Path(name).stem])
        return _assemble_context([source], fields)
    if suffix in (".xlsx", ".xlsm"):
        # openpyxl needs a file-like object.
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(content)
            tmp = Path(tf.name)
        try:
            fields, sheets, rows = _profile_xlsx(tmp)
        finally:
            tmp.unlink(missing_ok=True)
        source = DataSource(name=name, kind="url", location=url, row_count=rows, sheet_names=sheets)
        return _assemble_context([source], fields)
    raise ValueError(f"Unsupported remote file type for {url!r} (suffix={suffix!r})")


# ─────────────────────────────────────────────────────────────────────────────
# Description-based inference (no file)
# ─────────────────────────────────────────────────────────────────────────────

# Common field words users mention; mapped to a role/type when spotted.
_KNOWN_MEASURE_WORDS = [
    "revenue", "arr", "mrr", "cost", "profit", "margin", "budget", "ltv",
    "quantity", "units", "orders", "spend", "income", "sales", "amount", "price",
    "churn rate", "retention rate", "growth", "count",
]
_KNOWN_DATE_WORDS = ["date", "invoice date", "churn date", "order date", "period", "month", "year", "quarter"]


def context_from_description(description: str, *, entity_hint: str = "Data") -> DataContext:
    """Heuristically infer a DataContext from a prose description.

    Extracts comma/'and'-separated field-like tokens and classifies them by the
    same name heuristics used for columns. Records assumptions and a modest
    confidence because prose is inherently less certain than a real file.
    """
    text = description.strip()
    # Pull candidate field tokens: split on commas / "and" / newlines.
    raw_tokens = re.split(r",|\band\b|\n|;", text)
    tokens: list[str] = []
    for t in raw_tokens:
        t = t.strip().rstrip(".")
        # keep short noun-ish phrases (1-3 words), drop long sentences
        if 0 < len(t.split()) <= 3 and len(t) <= 40 and not t.lower().startswith(("i ", "build", "i want", "the ", "a ", "an ", "with", "for")):
            tokens.append(t)

    fields: list[FieldProfile] = []
    seen: set[str] = set()
    for tok in tokens:
        name = tok.strip().title()
        key = name.lower()
        if not name or key in seen:
            continue
        # Only accept tokens that look like data fields (contain a known word or
        # are single short nouns).
        looks_field = (
            any(w in key for w in _KNOWN_MEASURE_WORDS)
            or any(w in key for w in _KNOWN_DATE_WORDS)
            or (len(name.split()) <= 2 and key.isascii() and re.match(r"^[a-z][a-z /]*$", key))
        )
        if not looks_field:
            continue
        seen.add(key)
        ftype = _infer_type(name, [])
        role = _infer_role(name, ftype)
        fields.append(FieldProfile(name=name, entity=entity_hint, field_type=ftype, role=role))

    source = DataSource(name="user_description", kind="description", location="")
    ctx = _assemble_context([source], fields)
    ctx.assumptions.insert(0, "Data context inferred from a written description (no file provided); "
                             "field types are best-effort guesses.")
    # Prose confidence is capped lower than file-based profiling.
    ctx.confidence = round(min(ctx.confidence, 0.6) if fields else 0.35, 2)
    return ctx
